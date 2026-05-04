"""Extract a clean JSON dataset from the QuantumTech xlsx for visualizations."""
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).parent
SRC = ROOT / "QuantumTech HotSpot Investment Universe Database  (1).xlsx"
OUT = ROOT / "data" / "quantum_ecosystem.json"
OUT.parent.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(SRC, data_only=True)


def rows(sheet_name):
    ws = wb[sheet_name]
    headers = [
        (ws.cell(1, c).value or f"col{c}").strip() if isinstance(ws.cell(1, c).value, str)
        else (ws.cell(1, c).value or f"col{c}")
        for c in range(1, ws.max_column + 1)
    ]
    out = []
    for r in range(2, ws.max_row + 1):
        row = {}
        empty = True
        for i, h in enumerate(headers):
            v = ws.cell(r, i + 1).value
            if v not in (None, ""):
                empty = False
            row[h] = v
        if not empty:
            out.append(row)
    return out


def split_list(v):
    if not v or not isinstance(v, str):
        return []
    return [x.strip() for x in re.split(r"[;,]", v) if x.strip()]


# Vertical-specific company sheets to derive segment labels
VERTICAL_SHEETS = {
    "Quantum Bio & Medicine": "Quantum Bio & Medicine ",
    "Quantum Materials & NanoTech": "Quantum Materials & NanoTech",
    "Quantum AI & ML": "Quantum AI & ML",
    "Quantum Computing": "Quantum Computing HotSpot Compa",
    "Cyber Security": "Cyber Security Companies",
    "Communications": "Communications Companies",
}

vertical_by_company = {}
for vert, sheet in VERTICAL_SHEETS.items():
    if sheet in wb.sheetnames:
        for r in rows(sheet):
            name = r.get("name") or r.get("Name") or r.get("company") or r.get("Company")
            if name:
                vertical_by_company[name.strip()] = vert

# Master companies sheet
companies_raw = rows("Companies")
companies = []
for r in companies_raw:
    name = r.get("name")
    if not name:
        continue
    name = str(name).strip()
    investors = split_list(r.get("investors"))
    founders = split_list(r.get("founders"))
    categories = split_list(r.get("categories"))
    company = {
        "id": name,
        "name": name,
        "vertical": vertical_by_company.get(name, "Other"),
        "description": (r.get("short description") or r.get("description") or "")[:280] if isinstance(
            r.get("short description") or r.get("description"), str) else "",
        "founded": int(r["founded on"]) if isinstance(r.get("founded on"), (int, float)) else None,
        "country": (r.get("country") or "").strip() if isinstance(r.get("country"), str) else None,
        "region": (r.get("region") or "").strip() if isinstance(r.get("region"), str) else None,
        "city": (r.get("city") or "").strip() if isinstance(r.get("city"), str) else None,
        "employees": r.get("num_employees_enum"),
        "website": r.get("website"),
        "linkedin": r.get("linkedin"),
        "categories": categories,
        "investors": investors,
        "founders": founders,
    }
    companies.append(company)

# Investors — use Industry HotSpot - Investors sheet which has the totals
investors_raw = rows("Industry HotSpot - Investors")
investors = []
for r in investors_raw:
    name = r.get("Investor Name") or r.get("name") or r.get("Name")
    if not name:
        continue
    name = str(name).strip()
    investors.append({
        "id": name,
        "name": name,
        "type": r.get("Investor Type") or r.get("type"),
        "country": r.get("Investor Country") or r.get("country"),
        "city": r.get("Investor City") or r.get("city"),
        "investments_count": r.get("Number of Investments"),
        "lead_count": r.get("Number of Lead Investments"),
        "description": (r.get("Description") or "")[:280] if isinstance(r.get("Description"), str) else "",
    })

# Build the investor->company edges from the company.investors lists
company_names = {c["id"] for c in companies}
investor_names = {i["id"] for i in investors}

# Add any investors mentioned in companies.investors but not in investors sheet (so the graph is complete)
existing = set(investor_names)
for c in companies:
    for inv in c["investors"]:
        if inv not in existing:
            investors.append({
                "id": inv, "name": inv, "type": None, "country": None,
                "city": None, "investments_count": None, "lead_count": None,
                "description": "",
            })
            existing.add(inv)

investments = []
for c in companies:
    for inv in c["investors"]:
        investments.append({"investor": inv, "company": c["id"]})

# Founders graph — build founder nodes
founders_set = defaultdict(list)  # founder -> list of (company)
for c in companies:
    for f in c["founders"]:
        founders_set[f].append(c["id"])

founders = [
    {"id": f, "name": f, "companies": comps}
    for f, comps in founders_set.items()
]

# People — pull from UDM-style sheets that have Name/Company/Title/etc
PEOPLE_SHEETS = [
    "Companies UDMs (ALL in ONE)",
    "Sheet96", "Sheet97", "Sheet98", "Sheet99",
    "Investors UDMs",
    "Quantum Bio & Medicine UDMs",
    "Quantum Materials & NanoTech UD",
    "Quantum AI & ML UDMs",
    "Quantum AI & ML investors UDMs",
    "Cyber Security Companies UDMs",
    "Cyber Security Investors UDMs",
    "Communications Companies UDMs",
    "Communications Investors UDMs",
]

people_by_id = {}
for s in PEOPLE_SHEETS:
    if s not in wb.sheetnames:
        continue
    for r in rows(s):
        name = r.get("Name") or r.get("name")
        company = r.get("Company") or r.get("company")
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        if not name:
            continue
        title = r.get("Title") or r.get("title") or r.get("Headline")
        country = r.get("Country") or r.get("country")
        location = r.get("Location") or r.get("location")
        followers = r.get("Followers") or r.get("followers")
        linkedin = r.get("LinkedIn") or r.get("linkedin")
        key = (name, (company or "").strip() if isinstance(company, str) else "")
        if key in people_by_id:
            continue
        people_by_id[key] = {
            "id": f"{name}@{key[1]}" if key[1] else name,
            "name": name,
            "company": key[1] or None,
            "title": title.strip() if isinstance(title, str) else None,
            "country": country.strip() if isinstance(country, str) else None,
            "location": location.strip() if isinstance(location, str) else None,
            "followers": int(followers) if isinstance(followers, (int, float)) else None,
            "linkedin": linkedin if isinstance(linkedin, str) else None,
        }
people = list(people_by_id.values())

# City-level geographic aggregation for globe layer
geo = defaultdict(lambda: {"city": None, "country": None, "companies": [], "verticals": defaultdict(int)})
for c in companies:
    if c["city"] and c["country"]:
        key = f"{c['city']}|{c['country']}"
        g = geo[key]
        g["city"] = c["city"]
        g["country"] = c["country"]
        g["companies"].append(c["id"])
        g["verticals"][c["vertical"]] += 1

geo_list = [
    {
        "city": v["city"],
        "country": v["country"],
        "companies": v["companies"],
        "company_count": len(v["companies"]),
        "verticals": dict(v["verticals"]),
    }
    for v in geo.values()
]
geo_list.sort(key=lambda x: -x["company_count"])

# Co-investment edges (investors that share at least 2 portfolio companies)
inv_portfolio = defaultdict(set)
for e in investments:
    inv_portfolio[e["investor"]].add(e["company"])

co_invest = []
inv_list_sorted = sorted(inv_portfolio.keys())
for i, a in enumerate(inv_list_sorted):
    for b in inv_list_sorted[i + 1:]:
        shared = inv_portfolio[a] & inv_portfolio[b]
        if len(shared) >= 2:
            co_invest.append({"a": a, "b": b, "shared": len(shared)})

# Co-founding (founders sharing companies — minor) and shared-investor company pairs
shared_inv_pairs = defaultdict(int)
for inv, comps in inv_portfolio.items():
    comps = sorted(comps)
    for i, a in enumerate(comps):
        for b in comps[i + 1:]:
            shared_inv_pairs[(a, b)] += 1
co_company = [
    {"a": a, "b": b, "shared_investors": n}
    for (a, b), n in shared_inv_pairs.items() if n >= 2
]

dataset = {
    "meta": {
        "source": "QuantumTech HotSpot Investment Universe Database (1).xlsx",
        "verticals": list(VERTICAL_SHEETS.keys()),
        "company_count": len(companies),
        "investor_count": len(investors),
        "founder_count": len(founders),
        "investment_edges": len(investments),
        "city_count": len(geo_list),
    },
    "companies": companies,
    "investors": investors,
    "founders": founders,
    "people": people,
    "investments": investments,
    "co_invest": co_invest,
    "co_company": co_company,
    "geo": geo_list,
}
dataset["meta"]["people_count"] = len(people)

with open(OUT, "w") as f:
    json.dump(dataset, f, indent=None, default=str)

print("Wrote", OUT)
print("  companies:", len(companies))
print("  investors:", len(investors))
print("  founders:", len(founders))
print("  investments:", len(investments))
print("  co_invest:", len(co_invest))
print("  co_company pairs:", len(co_company))
print("  cities:", len(geo_list))
print("  people:", len(people))
